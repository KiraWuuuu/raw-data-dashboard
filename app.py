import io
import re
from typing import Optional, List

import numpy as np
import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill


# =========================
# 页面设置
# =========================
st.set_page_config(page_title="Prada Tracking Builder", layout="wide")
st.title("Prada Tracking Builder")
st.caption("上传 Preview Raw Data + Tracking Template + Mapping Config + Media Plan，自动生成输出 Excel。")


# =========================
# 标准字段（Campaign 页按 V1 结构输出）
# =========================
STANDARD_COLUMNS = [
    "Media", "Position", "Market", "Date",
    "Cost", "IMP", "CLICK", "CPM", "CPC", "CTR",
    "ENG", "Like", "Forward", "Comment", "Revenue", "Orders",
    "Landing page", "Campaign",
    "source_file", "target_sheet"
]


# =========================
# 通用工具函数
# =========================
def norm_text(v) -> str:
    if v is None:
        return ""
    s = str(v).strip().lower()
    s = s.replace("_", " ").replace("-", " ")
    s = re.sub(r"\s+", " ", s)
    return s


def clean_colnames(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df.columns = [str(c).strip() for c in df.columns]
    return df


def safe_get_col(df: pd.DataFrame, target_name: str) -> Optional[pd.Series]:
    target_norm = norm_text(target_name)
    for col in df.columns:
        if norm_text(col) == target_norm:
            return df[col]
    return None


def get_col_or_nan(df: pd.DataFrame, target_name: str) -> pd.Series:
    col = safe_get_col(df, target_name)
    if col is None:
        return pd.Series([np.nan] * len(df), index=df.index)
    return col


def to_number(v):
    if pd.isna(v) or v == "":
        return np.nan
    if isinstance(v, (int, float, np.integer, np.floating)):
        return float(v)

    s = str(v).strip()
    s = s.replace(",", "").replace("¥", "").replace("￥", "").replace("$", "")
    if s.endswith("%"):
        try:
            return float(s[:-1]) / 100
        except Exception:
            return np.nan

    try:
        return float(s)
    except Exception:
        return np.nan


def safe_div(a, b):
    if pd.isna(a) or pd.isna(b) or b == 0:
        return np.nan
    return a / b


def add_kpis(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()

    if "CTR" not in df.columns:
        df["CTR"] = np.nan
    if "CPM" not in df.columns:
        df["CPM"] = np.nan
    if "CPC" not in df.columns:
        df["CPC"] = np.nan

    df["CTR"] = df["CTR"].where(df["CTR"].notna(), df.apply(lambda x: safe_div(x["CLICK"], x["IMP"]), axis=1))
    df["CPM"] = df["CPM"].where(df["CPM"].notna(), df.apply(lambda x: safe_div(x["Cost"] * 1000, x["IMP"]), axis=1))
    df["CPC"] = df["CPC"].where(df["CPC"].notna(), df.apply(lambda x: safe_div(x["Cost"], x["CLICK"]), axis=1))

    return df


def map_market_by_keyword(text: str, market_map_df: pd.DataFrame) -> str:
    txt = str(text)
    mm = market_map_df.copy()

    if "priority" in mm.columns:
        mm["priority_num"] = pd.to_numeric(mm["priority"], errors="coerce")
        mm = mm.sort_values(by="priority_num", ascending=True, na_position="last")

    for _, row in mm.iterrows():
        keyword = str(row["keyword"])
        bucket = str(row["market_bucket"])
        if keyword and keyword in txt:
            return bucket

    return "Unknown"


def normalize_media_for_match(text: str) -> str:
    t = norm_text(text)
    if "微博" in t or "weibo" in t:
        return "weibo"
    if "抖音" in t or "douyin" in t:
        return "douyin"
    if "哔哩哔哩" in t or "bilibili" in t or "b站" in t:
        return "bilibili"
    if "微信" in t or "wechat" in t:
        return "wechat"
    return t


def normalize_position_for_match(text: str) -> str:
    t = norm_text(text)
    # 你可以继续扩展这层标准化
    if "开屏" in t or "opening" in t:
        return "opening"
    if "信息流" in t or "feed" in t or "feeds" in t:
        return "feeds"
    return t


# =========================
# 读取 Raw 文件
# =========================
def read_first_sheet(uploaded_file) -> pd.DataFrame:
    if uploaded_file.name.lower().endswith(".csv"):
        df = pd.read_csv(uploaded_file)
        return clean_colnames(df)

    data = uploaded_file.getvalue()
    xls = pd.ExcelFile(io.BytesIO(data), engine="openpyxl")
    first_sheet = xls.sheet_names[0]
    df = pd.read_excel(io.BytesIO(data), sheet_name=first_sheet, engine="openpyxl")
    return clean_colnames(df)


# =========================
# 读取 mapping_config
# =========================
def load_mapping_config(uploaded_file):
    cfg = pd.read_excel(uploaded_file, sheet_name=None, engine="openpyxl")

    if "file_mapping" not in cfg:
        raise ValueError("mapping_config.xlsx 缺少 sheet：file_mapping")
    if "market_mapping" not in cfg:
        raise ValueError("mapping_config.xlsx 缺少 sheet：market_mapping")

    file_map = clean_colnames(cfg["file_mapping"])
    market_map = clean_colnames(cfg["market_mapping"])

    required_cols = [
        "source_file", "target_sheet", "media_name",
        "position_mode", "position_value",
        "landing_page_default",
        "cost_total",
        "market_mode", "market_value",
        "cost_mode"
    ]
    missing = [c for c in required_cols if c not in file_map.columns]
    if missing:
        raise ValueError(f"file_mapping 缺少字段：{missing}")

    return file_map, market_map


def find_rule(file_map_df: pd.DataFrame, source_file_name: str) -> pd.Series:
    matched = file_map_df[file_map_df["source_file"].astype(str) == str(source_file_name)]
    if matched.empty:
        raise ValueError(f"mapping_config 的 file_mapping 中找不到 source_file = {source_file_name}")
    return matched.iloc[0]


# =========================
# 各媒体解析器
# =========================
def process_wechat(df: pd.DataFrame, cfg: pd.Series, market_map_df: pd.DataFrame) -> pd.DataFrame:
    """
    微信 raw:
    日期 / 广告名称 / 花费 / 曝光次数 / 点击次数 / 点击率 / 点赞次数 / 分享次数 / 评论次数 / 下单金额 / 下单次数
    """
    out = pd.DataFrame(index=df.index)

    out["Date"] = get_col_or_nan(df, "日期")
    out["Campaign"] = get_col_or_nan(df, "广告名称")
    out["Cost"] = get_col_or_nan(df, "花费").apply(to_number)
    out["IMP"] = get_col_or_nan(df, "曝光次数").apply(to_number)
    out["CLICK"] = get_col_or_nan(df, "点击次数").apply(to_number)
    out["CTR"] = get_col_or_nan(df, "点击率").apply(to_number)

    out["Like"] = get_col_or_nan(df, "点赞次数").apply(to_number)
    out["Forward"] = get_col_or_nan(df, "分享次数").apply(to_number)
    out["Comment"] = get_col_or_nan(df, "评论次数").apply(to_number)

    out["Revenue"] = get_col_or_nan(df, "下单金额").apply(to_number)
    out["Orders"] = get_col_or_nan(df, "下单次数").apply(to_number)

    out["ENG"] = out[["Like", "Forward", "Comment"]].fillna(0).sum(axis=1)

    out["Media"] = cfg["media_name"]

    if str(cfg["position_mode"]).lower() == "fixed":
        out["Position"] = cfg["position_value"]
    else:
        out["Position"] = get_col_or_nan(df, str(cfg["position_value"]))

    mmode = str(cfg["market_mode"]).lower()
    if mmode == "keyword":
        out["Market"] = out["Campaign"].apply(lambda x: map_market_by_keyword(x, market_map_df))
    elif mmode == "fixed":
        out["Market"] = cfg["market_value"]
    elif mmode == "raw":
        out["Market"] = get_col_or_nan(df, str(cfg["market_value"]))
    else:
        out["Market"] = "Unknown"

    out["Landing page"] = cfg["landing_page_default"]
    out["source_file"] = cfg["source_file"]
    out["target_sheet"] = cfg["target_sheet"]

    out["Date"] = pd.to_datetime(out["Date"], errors="coerce")
    out = add_kpis(out)

    for c in STANDARD_COLUMNS:
        if c not in out.columns:
            out[c] = np.nan

    return out[STANDARD_COLUMNS]


def process_douyin(df: pd.DataFrame, cfg: pd.Series) -> pd.DataFrame:
    """
    Douyin feeds / opening raw:
    Region / Type / SPID / Website / Channel / Ad Placement / Campaign ID / CampaignName / Date / Impression / Click / CTR
    """
    out = pd.DataFrame(index=df.index)

    out["Date"] = get_col_or_nan(df, "Date")
    out["Campaign"] = get_col_or_nan(df, "CampaignName")
    out["IMP"] = get_col_or_nan(df, "Impression").apply(to_number)
    out["CLICK"] = get_col_or_nan(df, "Click").apply(to_number)
    out["CTR"] = get_col_or_nan(df, "CTR").apply(to_number)

    out["Media"] = cfg["media_name"]

    if str(cfg["position_mode"]).lower() == "fixed":
        out["Position"] = cfg["position_value"]
    else:
        out["Position"] = get_col_or_nan(df, str(cfg["position_value"]))

    mmode = str(cfg["market_mode"]).lower()
    if mmode == "raw":
        out["Market"] = get_col_or_nan(df, str(cfg["market_value"]))
    elif mmode == "fixed":
        out["Market"] = cfg["market_value"]
    else:
        out["Market"] = "Unknown"

    out["Landing page"] = cfg["landing_page_default"]

    # 先不分摊，后面如果上传排期文件再从 Package Cost 回填
    out["Cost"] = np.nan

    out["Like"] = np.nan
    out["Forward"] = np.nan
    out["Comment"] = np.nan
    out["ENG"] = np.nan
    out["Revenue"] = np.nan
    out["Orders"] = np.nan

    out["source_file"] = cfg["source_file"]
    out["target_sheet"] = cfg["target_sheet"]

    out["Date"] = pd.to_datetime(out["Date"], errors="coerce")
    out = add_kpis(out)

    for c in STANDARD_COLUMNS:
        if c not in out.columns:
            out[c] = np.nan

    return out[STANDARD_COLUMNS]


def process_weibo(df: pd.DataFrame, cfg: pd.Series) -> pd.DataFrame:
    """
    微博 raw:
    日期 / 点位 / PV / Click
    """
    out = pd.DataFrame(index=df.index)

    out["Date"] = get_col_or_nan(df, "日期")
    out["Position"] = get_col_or_nan(df, "点位")
    out["IMP"] = get_col_or_nan(df, "PV").apply(to_number)
    out["CLICK"] = get_col_or_nan(df, "Click").apply(to_number)
    out["CTR"] = out.apply(lambda x: safe_div(x["CLICK"], x["IMP"]), axis=1)

    out["Media"] = cfg["media_name"]

    mmode = str(cfg["market_mode"]).lower()
    if mmode == "fixed":
        out["Market"] = cfg["market_value"]
    elif mmode == "raw":
        out["Market"] = get_col_or_nan(df, str(cfg["market_value"]))
    else:
        out["Market"] = "Unknown"

    out["Landing page"] = cfg["landing_page_default"]
    out["Campaign"] = np.nan

    # 先不分摊，后面如果上传排期文件再从 Package Cost 回填
    out["Cost"] = np.nan
    out["Like"] = np.nan
    out["Forward"] = np.nan
    out["Comment"] = np.nan
    out["ENG"] = np.nan
    out["Revenue"] = np.nan
    out["Orders"] = np.nan

    out["source_file"] = cfg["source_file"]
    out["target_sheet"] = cfg["target_sheet"]

    out["Date"] = pd.to_datetime(out["Date"], errors="coerce")
    out = add_kpis(out)

    for c in STANDARD_COLUMNS:
        if c not in out.columns:
            out[c] = np.nan

    return out[STANDARD_COLUMNS]


def process_bilibili(df: pd.DataFrame, cfg: pd.Series) -> pd.DataFrame:
    """
    Bilibili raw（如果你后面加上 B站 raw，可复用与 Douyin 相近的结构）
    """
    out = pd.DataFrame(index=df.index)

    out["Date"] = get_col_or_nan(df, "Date")
    out["Campaign"] = get_col_or_nan(df, "CampaignName")
    out["IMP"] = get_col_or_nan(df, "Impression").apply(to_number)
    out["CLICK"] = get_col_or_nan(df, "Click").apply(to_number)
    out["CTR"] = get_col_or_nan(df, "CTR").apply(to_number)

    out["Media"] = cfg["media_name"]

    if str(cfg["position_mode"]).lower() == "fixed":
        out["Position"] = cfg["position_value"]
    else:
        out["Position"] = get_col_or_nan(df, str(cfg["position_value"]))

    mmode = str(cfg["market_mode"]).lower()
    if mmode == "raw":
        out["Market"] = get_col_or_nan(df, str(cfg["market_value"]))
    elif mmode == "fixed":
        out["Market"] = cfg["market_value"]
    else:
        out["Market"] = "Unknown"

    out["Landing page"] = cfg["landing_page_default"]

    out["Cost"] = np.nan
    out["Like"] = np.nan
    out["Forward"] = np.nan
    out["Comment"] = np.nan
    out["ENG"] = np.nan
    out["Revenue"] = np.nan
    out["Orders"] = np.nan

    out["source_file"] = cfg["source_file"]
    out["target_sheet"] = cfg["target_sheet"]

    out["Date"] = pd.to_datetime(out["Date"], errors="coerce")
    out = add_kpis(out)

    for c in STANDARD_COLUMNS:
        if c not in out.columns:
            out[c] = np.nan

    return out[STANDARD_COLUMNS]


# =========================
# 解析固定模板排期（读取 Package Cost）
# =========================
def extract_period_from_sheet(df_raw: pd.DataFrame):
    """
    从固定模板第一页里找类似 2026.04.28-2026.05.11 的 campaign period
    """
    pattern = re.compile(r"(\d{4})\d{2}\d{2}\s*\d{2}\d{2}")
    for _, row in df_raw.iterrows():
        for cell in row.tolist():
            txt = str(cell)
            m = pattern.search(txt)
            if m:
                y1, mo1, d1, y2, mo2, d2 = m.groups()
                start = pd.Timestamp(int(y1), int(mo1), int(d1))
                end = pd.Timestamp(int(y2), int(mo2), int(d2))
                return start, end
    return None, None


def read_plan_fixed_template(uploaded_file) -> pd.DataFrame:
    """
    读取固定模板的媒体排期文件：
    - 找到包含 Website / Ad Format / Package Cost 的英文表头行
    - 读取后从 Package Cost 取得总成本
    - 根据日期列（28,29,30,1...11）拆成按天成本
    """
    data = uploaded_file.getvalue()
    raw = pd.read_excel(io.BytesIO(data), sheet_name=0, engine="openpyxl", header=None)
    raw = raw.fillna("")

    start_date, end_date = extract_period_from_sheet(raw)

    header_row_idx = None
    for i in range(len(raw)):
        row_vals = [str(v).strip() for v in raw.iloc[i].tolist()]
        joined = " | ".join(row_vals)
        if ("Website" in joined or "网站" in joined) and ("Ad Format" in joined or "广告形式/位置" in joined) and ("Package Cost" in joined or "打包价" in joined):
            header_row_idx = i
            break

    if header_row_idx is None:
        raise ValueError("排期文件中未找到包含 Website / Ad Format / Package Cost 的表头行")

    header = [str(v).strip() for v in raw.iloc[header_row_idx].tolist()]
    df = raw.iloc[header_row_idx + 1:].copy()
    df.columns = header
    df = df.reset_index(drop=True)
    df = clean_colnames(df)

    # 兼容中英文列名
    website_col = "Website" if "Website" in df.columns else "网站"
    adformat_col = "Ad Format" if "Ad Format" in df.columns else "广告形式/位置"
    region_col = "Region" if "Region" in df.columns else ("区域" if "区域" in df.columns else None)
    package_col = "Package Cost" if "Package Cost" in df.columns else ("打包价" if "打包价" in df.columns else None)

    if package_col is None:
        raise ValueError("排期文件中未找到 Package Cost / 打包价 列")

    # 找日期列：表头是纯数字
    day_cols = [c for c in df.columns if str(c).strip().isdigit()]

    rows = []

    for _, r in df.iterrows():
        media = r.get(website_col, np.nan)
        placement = r.get(adformat_col, np.nan)
        region = r.get(region_col, np.nan) if region_col is not None else np.nan
        package_cost = to_number(r.get(package_col, np.nan))

        if pd.isna(media) or pd.isna(placement) or pd.isna(package_cost):
            continue

        active_days = []
        for c in day_cols:
            val = str(r.get(c, "")).strip()
            if val not in ["", "0", "0.0", "-", "nan"]:
                active_days.append(int(c))

        if not active_days:
            continue

        if start_date is None or end_date is None:
            # 如果未找到 campaign period，就简单按 2026-04/05 推断（与你当前示例排期一致）
            expanded_dates = []
            for d in active_days:
                if d >= 28:
                    expanded_dates.append(pd.Timestamp(2026, 4, d))
                else:
                    expanded_dates.append(pd.Timestamp(2026, 5, d))
        else:
            expanded_dates = []
            cur = start_date
            while cur <= end_date:
                if cur.day in active_days:
                    expanded_dates.append(cur)
                cur += pd.Timedelta(days=1)

        if not expanded_dates:
            continue

        daily_cost = package_cost / len(expanded_dates)

        for dt in expanded_dates:
            rows.append({
                "Plan_Media": media,
                "Plan_Position": placement,
                "Plan_Market": region,
                "Date": dt,
                "Plan_Cost": daily_cost
            })

    return pd.DataFrame(rows)


def attach_cost_from_plan(raw_df: pd.DataFrame, plan_df: pd.DataFrame) -> pd.DataFrame:
    """
    用排期里的 Package Cost（拆成日成本后）去补 raw 的 Cost
    匹配键：Date + Media + Position（标准化后）
    """
    if plan_df is None or plan_df.empty:
        return raw_df

    df = raw_df.copy()
    pp = plan_df.copy()

    df["Date"] = pd.to_datetime(df["Date"], errors="coerce")
    pp["Date"] = pd.to_datetime(pp["Date"], errors="coerce")

    df["Media_norm"] = df["Media"].apply(normalize_media_for_match)
    df["Position_norm"] = df["Position"].apply(normalize_position_for_match)

    pp["Plan_Media_norm"] = pp["Plan_Media"].apply(normalize_media_for_match)
    pp["Plan_Position_norm"] = pp["Plan_Position"].apply(normalize_position_for_match)

    merged = df.merge(
        pp,
        how="left",
        left_on=["Date", "Media_norm", "Position_norm"],
        right_on=["Date", "Plan_Media_norm", "Plan_Position_norm"]
    )

    # 只在 raw 原本没有 Cost 时，用排期成本补
    merged["Cost"] = merged["Cost"].where(merged["Cost"].notna(), merged["Plan_Cost"])

    merged = merged.drop(columns=[
        "Media_norm", "Position_norm",
        "Plan_Media", "Plan_Position", "Plan_Market",
        "Plan_Cost", "Plan_Media_norm", "Plan_Position_norm"
    ], errors="ignore")

    # 回填后重新算 KPI
    merged = add_kpis(merged)

    # 只保留标准列
    keep_cols = [c for c in STANDARD_COLUMNS if c in merged.columns]
    extra_cols = [c for c in merged.columns if c not in keep_cols]
    merged = merged[keep_cols + extra_cols]

    return merged


# =========================
# Campaign 汇总（保持 V1 结构）
# =========================
def build_campaign_table(all_std: pd.DataFrame) -> pd.DataFrame:
    df = all_std.copy()
    df["Date"] = pd.to_datetime(df["Date"], errors="coerce").dt.strftime("%Y-%m-%d")

    group_cols = ["Media", "Position", "Market", "Date"]
    value_cols = ["Cost", "IMP", "CLICK", "ENG", "Like", "Forward", "Comment", "Revenue", "Orders"]

    agg = df.groupby(group_cols, dropna=False)[value_cols].sum(min_count=1).reset_index()

    agg["CTR"] = agg.apply(lambda x: safe_div(x["CLICK"], x["IMP"]), axis=1)
    agg["CPM"] = agg.apply(lambda x: safe_div(x["Cost"] * 1000, x["IMP"]), axis=1)
    agg["CPC"] = agg.apply(lambda x: safe_div(x["Cost"], x["CLICK"]), axis=1)

    agg = agg[[
        "Media", "Position", "Market", "Date",
        "Cost", "IMP", "CLICK", "CPM", "CPC", "CTR",
        "ENG", "Like", "Forward", "Comment", "Revenue", "Orders"
    ]]
    return agg


# =========================
# 写模板（重建 sheet，避免 merged cell）
# =========================
def recreate_sheet(wb, sheet_name: str):
    if sheet_name in wb.sheetnames:
        old_ws = wb[sheet_name]
        idx = wb.sheetnames.index(sheet_name)
        wb.remove(old_ws)
        new_ws = wb.create_sheet(title=sheet_name, index=idx)
    else:
        new_ws = wb.create_sheet(title=sheet_name)
    return new_ws


def write_table(ws, df: pd.DataFrame):
    header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    bold_font = Font(bold=True)

    # 写表头
    for c_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=c_idx, value=col_name)
        cell.fill = header_fill
        cell.font = bold_font

    # 写数据
    for r_idx, row in enumerate(df.itertuples(index=False), start=2):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=None if pd.isna(val) else val)

    # ===== 数字格式 =====
    accounting_format = '¥#,##0.00'
    integer_format = '#,##0'
    decimal_2_format = '0.00'
    percent_format = '0.00%'

    format_map = {
        # 会计
        "Cost": accounting_format,
        "Revenue": accounting_format,
        "花费": accounting_format,
        "下单金额": accounting_format,

        # 千位分隔
        "IMP": integer_format,
        "CLICK": integer_format,
        "ENG": integer_format,
        "Like": integer_format,
        "Forward": integer_format,
        "Comment": integer_format,
        "Orders": integer_format,
        "Impression": integer_format,
        "Click": integer_format,
        "曝光次数": integer_format,
        "点击次数": integer_format,
        "点赞次数": integer_format,
        "分享次数": integer_format,
        "评论次数": integer_format,
        "下单次数": integer_format,

        # 两位小数
        "CPM": decimal_2_format,
        "CPC": decimal_2_format,

        # 百分比
        "CTR": percent_format,
        "点击率": percent_format,
    }

    headers = [ws.cell(row=1, column=i).value for i in range(1, ws.max_column + 1)]

    for col_idx, col_name in enumerate(headers, start=1):
        if col_name in format_map:
            fmt = format_map[col_name]
            for row in range(2, ws.max_row + 1):
                ws.cell(row=row, column=col_idx).number_format = fmt

    # 自动列宽
    for col_cells in ws.columns:
        letter = col_cells[0].column_letter
        max_len = 0
        for cell in col_cells:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 50)


def build_output_workbook(template_file, all_std: pd.DataFrame) -> bytes:
    wb = load_workbook(io.BytesIO(template_file.getvalue()))

    # detail sheets
    for target_sheet, sub_df in all_std.groupby("target_sheet", dropna=False):
        target_sheet = str(target_sheet)

        export_df = sub_df.copy()
        export_df["Date"] = pd.to_datetime(export_df["Date"], errors="coerce").dt.strftime("%Y-%m-%d")

        if target_sheet in ["【WeChat Moments】", "【WeChat Banner】"]:
            out = export_df[
                ["Market", "Date", "Campaign", "Cost", "IMP", "CLICK", "CTR",
                 "Like", "Forward", "Comment", "Revenue", "Orders"]
            ].copy()
            out.columns = [
                "Market", "日期", "广告名称", "花费", "曝光次数", "点击次数", "点击率",
                "点赞次数", "分享次数", "评论次数", "下单金额", "下单次数"
            ]
        else:
            out = export_df[
                ["Market", "Position", "Campaign", "Date", "IMP", "CLICK", "CTR"]
            ].copy()
            out.columns = [
                "Region", "Ad Placement", "CampaignName", "Date", "Impression", "Click", "CTR"
            ]

        ws = recreate_sheet(wb, target_sheet)
        write_table(ws, out)

    # Campaign
    campaign_df = build_campaign_table(all_std)
    ws_campaign = recreate_sheet(wb, "Campaign")
    write_table(ws_campaign, campaign_df)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


# =========================
# 页面上传区
# =========================
raw_files = st.file_uploader(
    "Upload Raw Data",
    type=["xlsx", "xls", "csv"],
    accept_multiple_files=True
)

template_file = st.file_uploader(
    "Upload Template",
    type=["xlsx", "xls"],
    accept_multiple_files=False
)

config_file = st.file_uploader(
    "Upload Mapping Config",
    type=["xlsx", "xls"],
    accept_multiple_files=False
)

plan_files = st.file_uploader(
    "Upload Media Plan (optional)",
    type=["xlsx", "xls"],
    accept_multiple_files=True
)


# =========================
# 主流程
# =========================
if st.button("Generate Dashboard"):
    if not raw_files:
        st.error("请先上传 Preview Raw Data。")
        st.stop()

    if template_file is None:
        st.error("请先上传 Tracking Template。")
        st.stop()

    if config_file is None:
        st.error("请先上传 Mapping Config。")
        st.stop()

    try:
        file_map_df, market_map_df = load_mapping_config(config_file)
    except Exception as e:
        st.error(f"读取 mapping_config 失败：{e}")
        st.stop()

    all_parts = []

    for f in raw_files:
        try:
            cfg = find_rule(file_map_df, f.name)
            raw_df = read_first_sheet(f)
            lower_name = f.name.lower()

            if "wechat" in lower_name or "微信" in lower_name:
                std_df = process_wechat(raw_df, cfg, market_map_df)
            elif "douyin" in lower_name and "feeds" in lower_name:
                std_df = process_douyin(raw_df, cfg)
            elif "douyin" in lower_name and "opening" in lower_name:
                std_df = process_douyin(raw_df, cfg)
            elif "weibo" in lower_name or "微博" in lower_name:
                std_df = process_weibo(raw_df, cfg)
            elif "bili" in lower_name or "bilibili" in lower_name or "b站" in lower_name:
                std_df = process_bilibili(raw_df, cfg)
            else:
                st.warning(f"文件 {f.name} 没有匹配到规则，已跳过。")
                continue

            all_parts.append(std_df)

        except Exception as e:
            st.error(f"处理文件 {f.name} 失败：{e}")
            st.stop()

    if not all_parts:
        st.error("没有生成可用数据。")
        st.stop()

    all_std = pd.concat(all_parts, ignore_index=True)

    # 如果上传了排期文件，则从 Package Cost 补 Cost
    if plan_files:
        plan_parts = []
        for pf in plan_files:
            try:
                plan_parts.append(read_plan_fixed_template(pf))
            except Exception as e:
                st.error(f"读取排期文件 {pf.name} 失败：{e}")
                st.stop()

        if plan_parts:
            full_plan = pd.concat(plan_parts, ignore_index=True)
            all_std = attach_cost_from_plan(all_std, full_plan)

    st.subheader("Standardized Preview")
    preview_df = all_std.copy()
    preview_df["Date"] = pd.to_datetime(preview_df["Date"], errors="coerce").dt.strftime("%Y-%m-%d")
    st.dataframe(preview_df, use_container_width=True)

    try:
        output_bytes = build_output_workbook(template_file, all_std)
    except Exception as e:
        st.error(f"写入模板失败：{e}")
        st.stop()

    st.success("生成成功，可以下载输出文件。")
    st.download_button(
        "Download Output Excel",
        data=output_bytes,
        file_name="Prada_tracking_output_final.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
